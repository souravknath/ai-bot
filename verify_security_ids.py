import sqlite3
import logging
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import requests
import os
import time
from dotenv import load_dotenv

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Load environment variables
load_dotenv()

class SecurityIDVerifier:
    def __init__(self):
        """Initialize the SecurityIDVerifier with database and API details"""
        self.db_name = 'stock_data.db'
        self.conn = None
        self.cursor = None
        self.api_base_url = "https://api.dhan.co/v2"
        
        # Get API key from environment variable
        self.api_key = os.getenv("DHAN_API_KEY")
        if not self.api_key:
            logging.warning("API key not found in environment variables. Set DHAN_API_KEY in .env file.")
        
        # Headers for API requests
        self.headers = {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "access-token": self.api_key
        }
    
    def connect_to_db(self):
        """Connect to the SQLite database"""
        try:
            self.conn = sqlite3.connect(self.db_name)
            self.cursor = self.conn.cursor()
            logging.info(f"Connected to database: {self.db_name}")
            return True
        except sqlite3.Error as e:
            logging.error(f"Database connection error: {e}")
            return False
    
    def close_db(self):
        """Close the database connection"""
        if self.conn:
            self.conn.close()
            logging.info("Database connection closed")
    
    def get_all_stocks(self):
        """Get all stocks from the database"""
        try:
            self.cursor.execute("SELECT id, security_id, exchange_segment, symbol, name, instrument FROM stocks")
            return self.cursor.fetchall()
        except sqlite3.Error as e:
            logging.error(f"Error retrieving stocks: {e}")
            return []
    
    def export_to_excel(self, stocks):
        """Export stocks to Excel file"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Security IDs"
        
        # Add headers
        headers = ["ID", "Security ID", "Exchange Segment", "Symbol", "Name", "Instrument", "Status"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).value = header
        
        # Add data
        for row_num, stock in enumerate(stocks, 2):
            for col_num, value in enumerate(stock, 1):
                ws.cell(row=row_num, column=col_num).value = value
        
        # Adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save file
        filename = f"security_ids_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)
        logging.info(f"Exported {len(stocks)} stocks to {filename}")
        return filename
    
    def test_security_ids(self, stocks):
        """Test each security ID with the API to find problems"""
        yesterday = datetime.now().strftime("%Y-%m-%d")
        one_week_ago = (datetime.now() - datetime.timedelta(days=7)).strftime("%Y-%m-%d")
        
        # Create a new workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Security ID Verification"
        
        # Add headers
        headers = ["ID", "Security ID", "Exchange Segment", "Symbol", "Name", "Instrument", "Status"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).value = header
        
        # Define fill colors
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
        
        # Test each stock and add to Excel
        for row_num, stock in enumerate(stocks, 2):
            stock_id, security_id, exchange_segment, symbol, name, instrument = stock
            
            # Copy stock data to Excel
            for col_num, value in enumerate(stock, 1):
                ws.cell(row=row_num, column=col_num).value = value
            
            # Test API call
            endpoint = f"{self.api_base_url}/charts/historical"
            payload = {
                "securityId": security_id,
                "exchangeSegment": exchange_segment,
                "instrument": instrument,
                "fromDate": one_week_ago,
                "toDate": yesterday,
                "oi": False
            }
            
            try:
                response = requests.post(endpoint, headers=self.headers, json=payload)
                if response.status_code == 200:
                    status = "OK"
                    ws.cell(row=row_num, column=7).value = status
                    ws.cell(row=row_num, column=7).fill = green_fill
                else:
                    status = f"Error: {response.status_code} - {response.text}"
                    ws.cell(row=row_num, column=7).value = status
                    ws.cell(row=row_num, column=7).fill = red_fill
                    logging.error(f"Error testing {symbol} (ID: {security_id}): {status}")
            except Exception as e:
                status = f"Exception: {str(e)}"
                ws.cell(row=row_num, column=7).value = status
                ws.cell(row=row_num, column=7).fill = red_fill
                logging.error(f"Exception testing {symbol} (ID: {security_id}): {e}")
            
            # Sleep to avoid rate limits
            time.sleep(0.5)
        
        # Save file
        filename = f"security_id_verification_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)
        logging.info(f"Saved verification results to {filename}")
        return filename

def main():
    """Main function to verify security IDs"""
    verifier = SecurityIDVerifier()
    
    # Connect to database
    if not verifier.connect_to_db():
        return
    
    # Get all stocks
    stocks = verifier.get_all_stocks()
    if not stocks:
        logging.error("No stocks found in database")
        verifier.close_db()
        return
    
    logging.info(f"Found {len(stocks)} stocks in database")
    
    # Export to Excel
    excel_file = verifier.export_to_excel(stocks)
    
    # Test security IDs if requested
    user_input = input("Would you like to test all security IDs with the API? (y/n): ")
    if user_input.lower() == 'y':
        verification_file = verifier.test_security_ids(stocks)
        print(f"Security ID verification complete. Results saved to {verification_file}")
        print("Please check the file to identify invalid security IDs.")
        print("Red cells indicate failures that need to be fixed.")
    
    verifier.close_db()
    print(f"Security IDs exported to {excel_file}")
    print("You can now edit this file to update any incorrect security IDs.")

if __name__ == "__main__":
    main()